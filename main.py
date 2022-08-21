#!/usr/bin/env python
# coding: utf-8

# If you're not sure about the libraries listed here, just download them first - I intend to explain them in detail later.

import setup

setup._setup()

from requests import get
#import pandas as pd
from pandas import set_option, DataFrame
#import openpyxl # specially required for .xlsx file reading
from openpyxl import load_workbook
#import numpy as np
from numpy import nan
from io import BytesIO # Reading the .xlsx file
from IPython.display import display, HTML # A fancy way to display the dataframes, for debugging purposes
from tkinter import Tk, Label, OptionMenu, StringVar, Button, messagebox # Our GUI
from re import sub

colors = {}
wb = None

root = None

def main():
    if not setup.internet_is_available():
        messagebox.showinfo('No internet connection', 'Please connect to the internet before using this application.')
        return
    set_option(
        'expand_frame_repr', False # This is done so that dataframe is all presented on a single line
    )
    # And the below commands are an extension of the command above
    set_option('display.max_rows', 500)
    set_option('display.max_columns', 500)
    set_option('display.width', 1000)
    # Now this is where things get interesting
    ssid = "1uNvjtRePKdA-zpRzf1q-Ay0I5-Rh7mtvoI5odHKPARk"
    url = "https://docs.google.com/spreadsheets/export?exportFormat=xlsx&id=" + ssid
    r = get(url)
    data = BytesIO(r.content)
    global wb
    wb = load_workbook(filename=data) # Open up the excel sheet
    #wb = wb.active # Mark it as an active sheet
    days = ['Monday','Tuesday','Wednesday','Thursday','Friday'] # The sheets we require for now
    sheet = wb[days[-1]] # Retrieve a single day's sheet for now
    enough = False # A failsafe to just collect the degrees
    global colors
    colors = {}
    for i in sheet.iter_rows():
        for j in i:
            if j.value == 'Room':
                enough = True
                break
            if j.value is not None and j.value[0] in ['B', 'R', 'M', 'P']: # Stupid case to handle for BS, MS, PhD, and repeat
                colors[j.value] = j.fill.start_color.index # Store the cell value and cell color as dictionary
        if enough:
            break
    
    # Let's ask the user what degree they're taking (not handling MS/PhD/Repeat cases for now)
    GUI_init()
    root.mainloop()

CS_Sections = ["CS-A", "CS-B", "CS-C", "CS-D", "CS-E", "CS-F", "CS-G", "CS-H", "CS-J", "CS-K", "CS-V", "CS-Y", "CS-Z"]
DS_Sections = ["DS-A", "DS-B", "DS-C", "DS-D", "DS-M", "DS-N", "DS-U"]
SE_Sections = ["SE-A", "SE-B", "SE-C", "SE-D", "SE-E", "SE-F", "SE-G", "SE-P", "SE-Q", "SE-R", "SE-S"]
AI_Sections = ["AI-A", "AI-B", "AI-C", "AI-D", "AI-J", "AI-K"]
CY_Sections = ["CY-A", "CY-B", "CY-C", "CY-D", "CY-T"]

def process():
    info = degree.get() + " (" + batch.get() + ")"
    if info not in list(colors.keys()): # Display a prompt message stating that the combination does not exist
        messagebox.showinfo('Does not exist', 'Please enter a valid degree/batch year combination.')
    else: # Continue on and process the data
        global root
        root.destroy()
        root = Tk()
        root.eval('tk::PlaceWindow . center')
        Label(text=info + "\n").pack()
        days = ['Monday','Tuesday','Wednesday','Thursday','Friday'] # The sheets we require for now
        # Next 2 variables are for locating positions of Room and Lab cell values in sheet
        roomInfoIdx = 0
        hitRoomInfo = False
        sec = section.get() # get value of section
        for i in days:
            # We will use dataframes for data manipulation and searching
            dfClasses = None # A dataframe for classes
            dfHex = None # A dataframe for hex values of cell colour, corresponding with the classes dataframe
            
            sheet = wb[i] # Retrieve a sheet for that particular day
            # Temporary dataframes being made here for manipulation
            dfClasses = DataFrame(
                    sheet.values
                ) # Create a dataframe for accessing values with ease
            hexValues = [] # Makeshift list to create a dataframe manually
            for j in sheet.iter_rows():
                hexValues.append([k.fill.start_color.index if k.fill.start_color.index not in ['00000000', 'FFFFFFFF', 0] else nan for k in j])
            dfHex = DataFrame(
                hexValues
            )
            rowIdx = 0
            for j in sheet.iter_rows(): # For each row
                separated = False
                for k in j: # For each column (basically for each value)
                    if not hitRoomInfo and k.value != 'Room':
                        continue
                    elif not hitRoomInfo and k.value == "Room":
                        roomInfoIdx = rowIdx
                        hitRoomInfo = True
                        break
                    else: # if hitRoomInfo == true, then we will gather the dataframes and break subsequently
                        # First, drop the NaN columns
                        dfHex.columns = dfClasses.iloc[roomInfoIdx]
                        dfClasses = dfClasses.dropna( # For courses, this is how we'll do it
                            axis=1, how='all'
                        ).replace(
                            nan, '', regex=True
                        )
                        dfClasses.columns = dfClasses.iloc[roomInfoIdx]
                        # Now drop up till that specific row in Dataframe
                        dfClasses = dfClasses.iloc[roomInfoIdx+1:]
                        dfHex = dfHex.iloc[roomInfoIdx+1:]
                        separated = True
                        break
                if separated:
                    break
                rowIdx += 1
            # Now using the dataframes, we'll search for the appropriate listings
            colorToSearch = colors[info] # Extract the colour value to search
            firstHit = False
            for j in range(roomInfoIdx+1, len(dfClasses)): # Number of rows
                for k in dfClasses.columns: # Number of columns
                    if k == "Room":
                        continue # No need to traverse this column
                    if k is None or k == '': # Failsafe to ignore unused columns
                        continue
                    if dfHex.at[j,k] == colorToSearch and sec in dfClasses.at[j,k]:
                        # If found, note down the information (as text) (using time and room)
                        if not firstHit:
                            firstHit = True
                            Label(text="On " + i + ":").pack()
                        className = dfClasses.at[j,k]
                        className = sub("[\(\[].*?[\)\]]", "", className)
                        tex = className + " in " + dfClasses.at[j,'Room'] + ", timings: " + k
                        Label(text=tex).pack()
            if firstHit:
                Label(text="").pack()
        Button(root, text="Go back", command=go_back).pack()
        root.mainloop()

degree, batch = None, None
section = None

def go_back():
    GUI_init()

dropdownOption = None

def GUI_init():
    global root
    if root is not None:
        root.destroy()
    root = Tk()
    root.eval('tk::PlaceWindow . center')
    # Set up a degree dropdown
    global degree
    global batch
    global section
    Label(text="Select your degree").pack()
    degree = StringVar(root)
    degree.set("BS CS")
    global dropdownOption
    OptionMenu(
        root, degree, "BS CS", "BS DS", "BS SE", "BS AI", "BS CY", command=update_options
    ).pack()
    # The next two thing are BS-Exclusive options for now
    # Set up a batch dropdown
    Label(text="Select your batch (year)").pack()
    batch = StringVar(root)
    batch.set("2022")
    OptionMenu(root, batch, "2022", "2021", "2020", "2019").pack()
    # Set up a section dropdown (this may NOT be BS-exclusive after all)
    Label(text="Select your section").pack()
    section = StringVar(root)
    section.set("CS-A")
    sections = [i for i in CS_Sections]
    dropdownOption = OptionMenu(root, section, *sections)
    dropdownOption.pack()
    Button(root, text="Find your Timetable", command=process).pack()

# I've had little time to comment on this trickery - I'll come back to it.
def update_options(code):
    global root
    global dropdownOption
    menu = dropdownOption.children['menu']
    menu.delete(0, 'end')
    if code == "BS CS":
        for i in CS_Sections:
            menu.add_command(label=i, command=lambda v=i: newSec.set(v))
        section.set(CS_Sections[0])
    elif code == "BS DS":
        for i in DS_Sections:
            menu.add_command(label=i, command=lambda v=i: section.set(v))
        section.set(DS_Sections[0])
    elif code == "BS SE":
        for i in SE_Sections:
            menu.add_command(label=i, command=lambda v=i: section.set(v))
        section.set(SE_Sections[0])
    elif code == "BS AI":
        for i in AI_Sections:
            menu.add_command(label=i, command=lambda v=i: section.set(v))
        section.set(AI_Sections[0])
    elif code == "BS CY":
        for i in CY_Sections:
            menu.add_command(label=i, command=lambda v=i: section.set(v))
        section.set(CY_Sections[0])
    

if __name__ == "__main__":
    main()
